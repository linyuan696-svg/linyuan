#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
遮住的簾* — HackMD 同步工具
================================
依照 _sync/HackMD同步對照表.xlsx 的內容：
  1. 處理「動作」欄（新增 / 更新 / 刪除）
  2. 依表格順序重新產生四本書的目錄筆記，並直接覆蓋到 HackMD
  3. 把新網址寫回表格、清空已處理的動作欄

使用前準備（只需做一次）
------------------------
1. 安裝套件：
       python -m pip install requests openpyxl
2. 設定 API Token（每次開新的命令提示字元都要設一次）：
       set HACKMD_API_TOKEN=你的token

執行
----
       python sync_hackmd.py

只想重新產生目錄、不處理任何動作時，可加參數：
       python sync_hackmd.py --index-only
"""

import os
import re
import sys
import time
import shutil
import datetime
import requests
from openpyxl import load_workbook

# ============ 設定區：如有搬動資料夾請修改這裡 ============

REPO_ROOT = r"C:\Users\RiskaChen\Documents\GitHub\linyuan"

SHEET_PATH = os.path.join(REPO_ROOT, "_sync", "HackMD同步對照表.xlsx")
BACKUP_DIR = os.path.join(REPO_ROOT, "_sync", "backup")
DELETED_DIR = os.path.join(REPO_ROOT, "_sync", "deleted")

BOOKS = ["FANFIC", "ORIGINAL", "Commission", "雜文"]

# 四本書的目錄筆記（發布網址）
BOOK_INDEX_URLS = {
    "FANFIC":     "https://hackmd.io/@Riska0813/rJqpMMGvGg",
    "ORIGINAL":   "https://hackmd.io/@Riska0813/HyucXzfwMg",
    "Commission": "https://hackmd.io/@Riska0813/HJyMVfMvfe",
    "雜文":       "https://hackmd.io/@Riska0813/Hyo7Effwfl",
}

# 章節之間是否插入分隔線
ADD_DIVIDER = True
DIVIDER_TEXT = "──────────"

# 保留幾份備份
KEEP_BACKUPS = 10

# API
API_BASE = "https://api.hackmd.io/v1"
SLEEP_SECONDS = 1.0
MAX_RETRIES = 6

COL = {"書": 1, "章節": 2, "分區": 3, "顯示標題": 4, "檔案路徑": 5, "動作": 6, "網址": 7}

# ============ 以下不需修改 ============

API_TOKEN = os.environ.get("HACKMD_API_TOKEN")
if not API_TOKEN:
    print("錯誤：找不到 HACKMD_API_TOKEN 環境變數。")
    print("請先執行：set HACKMD_API_TOKEN=你的token")
    sys.exit(1)

HEADERS = {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}


def api(method, path, **kwargs):
    """呼叫 API，遇到 429 自動等待重試"""
    url = f"{API_BASE}{path}"
    delay = 5
    for attempt in range(1, MAX_RETRIES + 1):
        resp = requests.request(method, url, headers=HEADERS, timeout=30, **kwargs)
        if resp.status_code != 429:
            return resp
        wait = float(resp.headers.get("Retry-After", delay))
        print(f"    ⚠ 被限速，等待 {wait:.0f} 秒後重試（{attempt}/{MAX_RETRIES}）")
        time.sleep(wait)
        delay = min(delay * 2, 60)
    return resp


def list_notes():
    resp = api("GET", "/notes")
    if resp.status_code != 200:
        raise RuntimeError(f"取得筆記列表失敗（{resp.status_code}）：{resp.text[:200]}")
    return resp.json()


def create_note(title, content):
    payload = {
        "title": title or "無標題",
        "content": content,
        "readPermission": "guest",
        "writePermission": "owner",
        "commentPermission": "everyone",
    }
    resp = api("POST", "/notes", json=payload)
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"建立失敗（{resp.status_code}）：{resp.text[:200]}")
    d = resp.json()
    return d["id"], d.get("publishLink") or f"https://hackmd.io/{d['id']}"


def update_note(note_id, content):
    resp = api("PATCH", f"/notes/{note_id}", json={"content": content})
    if resp.status_code not in (200, 202, 204):
        raise RuntimeError(f"更新失敗（{resp.status_code}）：{resp.text[:200]}")


def get_note(note_id):
    resp = api("GET", f"/notes/{note_id}")
    if resp.status_code != 200:
        return None
    return resp.json()


def delete_note(note_id):
    resp = api("DELETE", f"/notes/{note_id}")
    if resp.status_code not in (200, 202, 204):
        raise RuntimeError(f"刪除失敗（{resp.status_code}）：{resp.text[:200]}")


def backup_sheet():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_DIR, f"對照表_{stamp}.xlsx")
    shutil.copy2(SHEET_PATH, dst)
    print(f"已備份對照表 → {os.path.relpath(dst, REPO_ROOT)}")
    # 只保留最近幾份
    backups = sorted(
        (f for f in os.listdir(BACKUP_DIR) if f.startswith("對照表_")),
        reverse=True,
    )
    for old in backups[KEEP_BACKUPS:]:
        os.remove(os.path.join(BACKUP_DIR, old))


def read_rows(ws):
    """讀取一個分頁的所有資料列，回傳 [(excel列號, dict), ...]"""
    out = []
    for i in range(2, ws.max_row + 1):
        title = ws.cell(i, COL["顯示標題"]).value
        if not title or not str(title).strip():
            continue
        out.append((i, {
            "章節": (ws.cell(i, COL["章節"]).value or "").strip(),
            "分區": (ws.cell(i, COL["分區"]).value or "").strip(),
            "標題": str(title).strip(),
            "路徑": (ws.cell(i, COL["檔案路徑"]).value or "").strip(),
            "動作": (ws.cell(i, COL["動作"]).value or "").strip(),
            "網址": (ws.cell(i, COL["網址"]).value or "").strip(),
        }))
    return out


def build_index(book, rows):
    """依照表格順序組出目錄筆記內容"""
    lines = [book, "===", ""]
    chapter_order, chapters = [], {}
    for _, r in rows:
        ch = r["章節"] or "(未分章節)"
        if ch not in chapters:
            chapters[ch] = {"直屬": [], "分區順序": [], "分區": {}}
            chapter_order.append(ch)
        c = chapters[ch]
        if r["分區"]:
            if r["分區"] not in c["分區"]:
                c["分區"][r["分區"]] = []
                c["分區順序"].append(r["分區"])
            c["分區"][r["分區"]].append(r)
        else:
            c["直屬"].append(r)

    for idx, ch in enumerate(chapter_order):
        c = chapters[ch]
        lines.append(f"## {ch}")
        lines.append("")
        for r in c["直屬"]:
            lines.append(f"- [{r['標題']}]({r['網址']})")
        if c["直屬"]:
            lines.append("")
        for secname in c["分區順序"]:
            lines.append(f"### {secname} [close]")
            lines.append("---")
            for r in c["分區"][secname]:
                lines.append(f"- [{r['標題']}]({r['網址']})")
            lines.append("")
        if ADD_DIVIDER and idx < len(chapter_order) - 1:
            lines.append(f"### {DIVIDER_TEXT}")
            lines.append("---")
            lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def main():
    index_only = "--index-only" in sys.argv

    if not os.path.exists(SHEET_PATH):
        print(f"錯誤：找不到對照表 {SHEET_PATH}")
        sys.exit(1)

    print("讀取 HackMD 現有筆記…")
    notes = list_notes()
    by_link = {n.get("publishLink"): n for n in notes if n.get("publishLink")}
    print(f"目前帳號共有 {len(notes)} 篇筆記")

    backup_sheet()
    wb = load_workbook(SHEET_PATH)

    pending_delete = []   # (book, excel列號, row)
    to_create, to_update = [], []

    for book in BOOKS:
        if book not in wb.sheetnames:
            print(f"⚠ 找不到分頁「{book}」，略過")
            continue
        for i, r in read_rows(wb[book]):
            act = r["動作"]
            if not act or index_only:
                continue
            if act == "新增":
                to_create.append((book, i, r))
            elif act == "更新":
                to_update.append((book, i, r))
            elif act == "刪除":
                pending_delete.append((book, i, r))
            else:
                print(f"⚠ {book} 第{i}列：無法辨識的動作「{act}」，已略過")

    # ---------- 刪除（先確認）----------
    if pending_delete:
        print("\n" + "=" * 50)
        print(f"即將刪除 {len(pending_delete)} 篇筆記：")
        for book, i, r in pending_delete:
            print(f"  · [{book}] {r['標題']}")
        print("=" * 50)
        print("刪除後將從目錄與 HackMD 一併移除（內容會先備份到 _sync/deleted/）")
        ans = input("確定要刪除嗎？請輸入 yes 繼續：").strip().lower()
        if ans != "yes":
            print("已取消刪除，其他動作照常執行。")
            pending_delete = []

    if pending_delete:
        os.makedirs(DELETED_DIR, exist_ok=True)
        for book, i, r in list(pending_delete):
            note = by_link.get(r["網址"])
            if not note:
                print(f"  找不到對應筆記，略過刪除：{r['標題']}")
                pending_delete.remove((book, i, r))
                continue
            try:
                full = get_note(note["id"])
                if full and full.get("content"):
                    safe = re.sub(r'[\\/:*?"<>|]', "_", r["標題"])
                    stamp = datetime.datetime.now().strftime("%Y%m%d")
                    with open(os.path.join(DELETED_DIR, f"{stamp}_{book}_{safe}.md"),
                              "w", encoding="utf-8") as f:
                        f.write(full["content"])
                delete_note(note["id"])
                print(f"  已刪除：{r['標題']}")
            except Exception as e:
                print(f"  刪除失敗：{r['標題']} → {e}")
                pending_delete.remove((book, i, r))
            time.sleep(SLEEP_SECONDS)

    # ---------- 新增 ----------
    for book, i, r in to_create:
        path = os.path.join(REPO_ROOT, r["路徑"].replace("/", os.sep)) if r["路徑"] else None
        if not path or not os.path.exists(path):
            print(f"  ✗ 新增失敗（找不到檔案）：{r['標題']} ← {r['路徑']}")
            continue
        content = open(path, encoding="utf-8").read()
        try:
            _id, link = create_note(r["標題"], content)
            wb[book].cell(i, COL["網址"], link)
            wb[book].cell(i, COL["動作"], None)
            print(f"  ✓ 已新增：{r['標題']}")
        except Exception as e:
            print(f"  ✗ 新增失敗：{r['標題']} → {e}")
        time.sleep(SLEEP_SECONDS)

    # ---------- 更新 ----------
    for book, i, r in to_update:
        path = os.path.join(REPO_ROOT, r["路徑"].replace("/", os.sep)) if r["路徑"] else None
        if not path or not os.path.exists(path):
            print(f"  ✗ 更新失敗（找不到檔案）：{r['標題']} ← {r['路徑']}")
            continue
        note = by_link.get(r["網址"])
        if not note:
            print(f"  ✗ 更新失敗（找不到對應筆記）：{r['標題']}")
            continue
        try:
            update_note(note["id"], open(path, encoding="utf-8").read())
            wb[book].cell(i, COL["動作"], None)
            print(f"  ✓ 已更新：{r['標題']}")
        except Exception as e:
            print(f"  ✗ 更新失敗：{r['標題']} → {e}")
        time.sleep(SLEEP_SECONDS)

    # ---------- 從表格移除已刪除的列（由下往上）----------
    for book in BOOKS:
        del_rows = sorted([i for b, i, _ in pending_delete if b == book], reverse=True)
        for i in del_rows:
            wb[book].delete_rows(i)

    # ---------- 重新產生四本目錄並覆蓋 ----------
    print("\n重新產生目錄筆記…")
    for book in BOOKS:
        if book not in wb.sheetnames:
            continue
        rows = read_rows(wb[book])
        missing = [r["標題"] for _, r in rows if not r["網址"]]
        if missing:
            print(f"  ⚠ {book}：有 {len(missing)} 篇沒有網址，將不會出現在目錄")
            for t in missing[:5]:
                print(f"      · {t}")
        rows = [(i, r) for i, r in rows if r["網址"]]
        content = build_index(book, rows)

        target = by_link.get(BOOK_INDEX_URLS.get(book, ""))
        if not target:
            out = os.path.join(REPO_ROOT, "_sync", f"目錄_{book}.md")
            with open(out, "w", encoding="utf-8") as f:
                f.write(content)
            print(f"  ⚠ {book}：找不到目錄筆記，已存成檔案 {os.path.basename(out)}，請手動貼上")
            continue
        try:
            update_note(target["id"], content)
            print(f"  ✓ {book} 目錄已更新（{len(rows)} 篇）")
        except Exception as e:
            print(f"  ✗ {book} 目錄更新失敗：{e}")
        time.sleep(SLEEP_SECONDS)

    wb.save(SHEET_PATH)
    print("\n對照表已儲存。")
    print("提醒：記得用 GitHub Desktop 把 _sync 的變更 commit + push。")


if __name__ == "__main__":
    main()
